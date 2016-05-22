import os


def read(workbook_name):
    import openpyxl
    if os.path.exists(workbook_name):
        workbook = openpyxl.load_workbook(workbook_name)
    else:
        print("Excel File Not Found ! ")
        raise Exception
    worksheet = workbook.active
    read_data = [[0 for x in range(worksheet.max_column + 1)] for x in range(worksheet.max_row + 1)]
    j = 1
    for row in worksheet.rows:
        i = 1
        for cell in row:
            if cell.value is None:
                read_data[j][i] = ""
            else:
                read_data[j][i] = cell.value
            i += 1
        j += 1
    return read_data
