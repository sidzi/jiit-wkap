import openpyxl
import os


class ExcelWriter:
    def __init__(self, wb_name):
        if not (str(wb_name).endswith('.xlsx') or str(wb_name).endswith('.xls')):
            self.workbook_name = str(wb_name) + '.xlsx'
        if os.path.exists(wb_name):
            self.workbook = openpyxl.load_workbook(wb_name)
        else:
            self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

    def write(self, row, col, value):
        self.worksheet.cell(row=row, column=col, value=value)

    def close(self):
        self.workbook.save(filename=self.workbook_name)
